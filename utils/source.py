import os
os.environ['KMP_DUPLICATE_LIB_OK']='True'
import easyocr
import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
import tempfile
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import re


# Initialize the EasyOCR reader with the desired languages
reader = easyocr.Reader(['en'])

# Initialize an empty list to store results
results = []

# Initialize an empty list to store temporary image file paths
temp_image_paths = []

# List of characters to check for in the "Text" column
characters_to_check = ['Z', 'W', 'M', 'N', 'O', 'I', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')', ' ','z','w','n','o','i','+','-','*','/',]

# Function to handle the file selection, exlcuding any link for network
def select_directory():
    network_drive_regex = r"\/\/[a-zA-Z0-9.-]+\/[a-zA-Z0-9.-]+"
    imagdir = filedialog.askdirectory()
    if imagdir and not re.match(network_drive_regex, imagdir):
        # Existing code to handle the directory selection
        global image_dir
        image_dir = imagdir
        label.config(text=f"Selected Directory: {image_dir}")
        input_frame.pack()
    else:
        # Handling the case where a network path is selected
        label.config(text="Error: Network Path not allowed")


# Function to ask for the Excel file location
def select_output_file():
    global output_excel_file
    output_excel_file = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if output_excel_file:
        output_label.config(text=f"Excel Output File: {output_excel_file}")
        process_button.config(state=tk.NORMAL)

# Function to execute the OCR process
def process_images():
    if not 'image_dir' in globals():
        messagebox.showerror("Error", "Please select a directory first.")
        return

    results.clear()
    temp_image_paths.clear()
    label.config(text="Processing...")

    left = int(left_entry.get())
    top = int(top_entry.get())
    right = int(right_entry.get())
    bottom = int(bottom_entry.get())

    crop_coords = (left, top, right, bottom)

    # Iterate through the images in the selected directory
    for filename in os.listdir(image_dir):
        if filename.endswith('.jpg'):
            image_path = os.path.join(image_dir, filename)

            # Open the image using PIL
            img = Image.open(image_path)

            # Crop the image using the user-input coordinates
            cropped_img = img.crop(crop_coords)

            # Save the cropped image as a temporary file
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as temp_file:
                temp_image_path = temp_file.name
                cropped_img.save(temp_image_path)
                temp_image_paths.append(temp_image_path)

            # Read text from the temporary image file
            result = reader.readtext(temp_image_path, detail=0, paragraph=True)

            # Join the extracted text and convert to uppercase
            text = ' '.join(result).upper()

            # Initialize confidence level at 100%
            confidence_level = 100

            # Check for conditions and adjust the confidence level
            for char in characters_to_check:
                if char in text:
                    confidence_level -= 5

            if any(c.isdigit() for c in text) and sum(c.isdigit() for c in text) > 6:
                confidence_level -= 30

            if not any(c.isalpha() for c in text):
                confidence_level -= 20

            if sum(c.isalpha() for c in text) > 3:
                confidence_level -= 20

            results.append({'Image File': filename, 'Text': text, 'Confidence Level': confidence_level, 'Image Path': temp_image_path})

    # Create a DataFrame from the results
    df = pd.DataFrame(results)
    
    # Remove spaces from the 'Text' column
    df['Text'] = df['Text'].str.replace(' ', '')
    
    # Remove special characters from the 'Text' column
    df['Text'] = df['Text'].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x))
    
    # Duplicate the "Text" column and name the duplicate "Text2"
    df['Text2'] = df['Text'].copy()
    
    # Split the 'Text2' column into two new columns
    df['Series'] = df['Text2'].str.slice(0, -6)
    df['Serial'] = df['Text2'].str.slice(-6)
    
                                                        #Learn and edit here
    df['Series'] = df['Series'].str.replace('4', 'A')
    df['Series'] = df['Series'].str.replace('6', 'G')
    df['Series'] = df['Series'].str.replace('0', 'O')
    df['Series'] = df['Series'].str.replace('2', 'Z')
    
    df['Serial'] = df['Serial'].str.replace('O', '0')
    df['Serial'] = df['Serial'].str.replace('l', '1')
    df['Serial'] = df['Serial'].str.replace('i', '1')
    df['Serial'] = df['Serial'].str.replace('I', '1')
    
    
    # Combine 'Series' and 'Serial' columns into a new column 'Combined'
    df['PB35_Listing'] = df['Series'] + df['Serial']
        
    print(df)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the Excel file
    ws['A1'] = 'Image File'
    #ws['B1'] = 'Text'
    ws['C1'] = 'Confidence Level'
    ws['D1'] = 'Series'
    ws['E1'] = 'Serial'
    ws['F1'] = 'PB35_Listing'

    # Define cell fill colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Paste the data and resized images into the Excel file
    for row, data in enumerate(df.itertuples(), start=2):
        image_file = data[1]
        text = data[2]
        confidence_level = data[3]
        temp_image_path = data[4]
        text2 = data[5]  # Assuming 'Text2' is the last column in the DataFrame
        series = data[6]  # Assuming 'Series' is the second-to-last column
        serial = data[7]  # Assuming 'Serial' is the third-to-last column
        pb35_listing = data[8]
        
        ws.cell(row=row, column=1, value=image_file)
        #ws.cell(row=row, column=2, value=text)
        ws.cell(row=row, column=3, value=confidence_level)
        ws.cell(row=row, column=4, value=series)
        ws.cell(row=row, column=5, value=serial)
        ws.cell(row=row, column=6, value=pb35_listing)

        # Apply cell fill colors based on confidence level
        if confidence_level < 100:
            ws.cell(row=row, column=3).fill = yellow_fill
        if confidence_level < 90:
            ws.cell(row=row, column=3).fill = orange_fill
            
        # Apply cell fill colors based on conditions for the 'Series' column
        if any(char.isdigit() for char in series):
            ws.cell(row=row, column=4).fill = yellow_fill
        if sum(char.isalpha() for char in series) > 3:
            ws.cell(row=row, column=4).fill = orange_fill
            
            
        # Apply cell fill colors based on conditions for the 'Serial' column
        if any(char.isalpha() for char in serial):
            ws.cell(row=row, column=5).fill = yellow_fill
        if len([char for char in serial if char.isdigit()]) > 6:
            ws.cell(row=row, column=5).fill = yellow_fill
                       

        # Add the image to the Excel file
        excel_image = ExcelImage(temp_image_path)
        excel_image.width = int(excel_image.width * (1/5))
        excel_image.height = int(excel_image.height * (1/5))
        ws.add_image(excel_image, f'H{row}')

    # Save the Excel file
    wb.save(output_excel_file)

    label.config(text=f"Text extracted from images in '{image_dir}' and saved to '{output_excel_file}'.")

# Create a Tkinter window with a dark theme
window = tk.Tk()
window.title("Text Extraction for PB35 by KARACHI OFFICE")

# Set the background color to black
window.configure(bg="black")

# Create a label for the selected directory with white text
label = tk.Label(window, text="                          Select a directory with Prize Bond JPEG images                       ", bg="black", fg="white")
label.pack(padx=10, pady=10)

# Set the background color of the "Select Directory" button to dark gray
select_button = tk.Button(window, text="Select Directory", command=select_directory, bg="dark gray")
select_button.pack(pady=5)

# Create a frame for input fields with a black background
input_frame = tk.Frame(window, bg="black")

# Create input fields for cropping coordinates with white text
left_label = tk.Label(input_frame, text="Left:", bg="black", fg="white")
left_label.pack(side="left")
left_entry = tk.Entry(input_frame)
left_entry.pack(side="left", padx=5)

top_label = tk.Label(input_frame, text="Top:", bg="black", fg="white")
top_label.pack(side="left")
top_entry = tk.Entry(input_frame)
top_entry.pack(side="left", padx=5)

right_label = tk.Label(input_frame, text="Right:", bg="black", fg="white")
right_label.pack(side="left")
right_entry = tk.Entry(input_frame)
right_entry.pack(side="left", padx=5)

bottom_label = tk.Label(input_frame, text="Bottom:", bg="black", fg="white")
bottom_label.pack(side="left")
bottom_entry = tk.Entry(input_frame)
bottom_entry.pack(side="left", padx=5)

# Create a label for the Excel output file selection with white text
output_label = tk.Label(window, text="Select an Excel output file location", bg="black", fg="white")
output_label.pack(padx=10, pady=5)

# Set the background color of the "Select Excel Output File" button to dark gray
select_output_button = tk.Button(window, text="Select Excel Output File", command=select_output_file, bg="dark gray")
select_output_button.pack()

# Create a button to process the images (initially disabled)
# Set the background color to dark green
process_button = tk.Button(input_frame, text="Process Images", command=process_images, state=tk.DISABLED, bg="dark gray")
process_button.pack(pady=10)

# Hide the input frame initially
input_frame.pack_forget()

window.mainloop()
